from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
import pandas as pd

from excel_reader import FORM_COLUMNS, BOT_COLUMNS, ALL_COLUMNS, SHEET_NAME


# ── Styling constants (match the input template exactly) ─────────────────────
FONT_NAME   = "Arial"
FONT_SIZE   = 10

HDR_BG      = "4472C4"          # blue header
HDR_FG      = "FFFFFF"
HDR_HEIGHT  = 32

STATUS_FILLS = {
    "SUCCESS": PatternFill("solid", start_color="C6EFCE"),  # green
    "FAILED":  PatternFill("solid", start_color="FFC7CE"),  # red
}

SUMMARY_HDR_BG = "ED7D31"       # orange — matches Valid Options sheet

COL_WIDTHS = {
    "theme":                         40,
    "activity_name":                 45,
    "focus_area":                    32,
    "activity_type":                 20,
    "activity_description":          45,
    "pdi_indicator":                 50,
    "activity_for":                  12,
    "targeted_populace":             30,
    "activity_nature":               18,
    "is_directly_funded_by_panchayat": 14,
    "estimated_completion_year":     10,
    "estimated_completion_month":    10,
    "estimated_completion_days":     10,
    "start_year":                    12,
    "start_month":                   14,
    "expected_beneficiary_general":  12,
    "expected_beneficiary_sc":        9,
    "expected_beneficiary_st":        9,
    "estimated_total_cost":          16,
    "status":                        12,
    "processed_time":                20,
    "error_message":                 60,
}

_thin  = Side(style="thin",   color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr_cell(ws, row: int, col: int, value: str,
              bg: str = HDR_BG, fg: str = HDR_FG) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=fg, name=FONT_NAME, size=FONT_SIZE)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _BORDER


def _normalize_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if pd.isna(value):
        return ""
    return str(value)


def _data_cell(ws, row: int, col: int, value: object,
               fill: PatternFill | None = None) -> None:
    c = ws.cell(row=row, column=col, value=_normalize_cell_value(value))
    c.font      = Font(name=FONT_NAME, size=FONT_SIZE)
    c.alignment = Alignment(vertical="top", wrap_text=False)
    c.border    = _BORDER
    if fill:
        c.fill = fill


class ResultWriter:
    """
    Keeps the full DataFrame in memory; flushes to a styled .xlsx after every row.
    Output format exactly matches the input template (eGram Activity Data sheet
    + Automation_Summary sheet).
    Only the 21 form columns + 3 bot columns are written — no extras.
    """

    def __init__(self, input_path: str, output_path: str) -> None:
        self.input_path  = Path(input_path)
        self.output_path = Path(output_path)
        self._df: pd.DataFrame | None = None

    # ── In-memory DataFrame ───────────────────────────────────────────────────

    def _get_df(self) -> pd.DataFrame:
        if self._df is None:
            # Base the output on the input workbook so the actual form values are preserved.
            base_df = pd.read_excel(
                self.input_path,
                sheet_name=SHEET_NAME,
                dtype=str,
                keep_default_na=False,
            )
            base_df.columns = [c.strip().lower().replace(" ", "_") for c in base_df.columns]
            base_df = base_df[[c for c in base_df.columns if c in ALL_COLUMNS]]
            for col in ALL_COLUMNS:
                if col not in base_df.columns:
                    base_df[col] = ""
            base_df = base_df[ALL_COLUMNS]

            # If an output workbook already exists, preserve any bot status columns from it.
            if self.output_path.exists():
                try:
                    prev_df = pd.read_excel(
                        self.output_path,
                        sheet_name=SHEET_NAME,
                        dtype=str,
                        keep_default_na=False,
                    )
                    prev_df.columns = [c.strip().lower().replace(" ", "_") for c in prev_df.columns]
                    prev_df = prev_df[[c for c in prev_df.columns if c in ALL_COLUMNS]]
                    for col in ALL_COLUMNS:
                        if col not in prev_df.columns:
                            prev_df[col] = ""
                    prev_df = prev_df[ALL_COLUMNS]

                    for col in BOT_COLUMNS:
                        if col in prev_df.columns:
                            base_df[col] = prev_df[col].reset_index(drop=True).reindex(base_df.index, fill_value="")
                except Exception:
                    pass

            self._df = base_df
        return self._df

    # ── Per-row update (O(1) in memory) ──────────────────────────────────────

    def write_result(
        self,
        row_index: int,
        status: str,
        error_message: str = "",
        application_number: str = "",   # reserved, not written to sheet
        transaction_id: str = "",       # reserved, not written to sheet
    ) -> None:
        df = self._get_df()
        df.at[row_index, "status"]         = status
        df.at[row_index, "processed_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df.at[row_index, "error_message"]  = error_message
        self._flush()

    # ── Styled flush to disk ──────────────────────────────────────────────────

    def _flush(self) -> None:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        df = self._get_df()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.row_dimensions[1].height = HDR_HEIGHT

        # Header row
        for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
            _hdr_cell(ws, 1, col_idx, col_name)
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)

        # Data rows
        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            status_val = str(row.get("status", "")).strip().upper()
            row_fill   = STATUS_FILLS.get(status_val)
            for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
                val = row.get(col_name, "")
                # Status cell always gets colour; other cells get it only on success/fail rows
                cell_fill = row_fill if col_name in ("status", "processed_time", "error_message") else None
                _data_cell(ws, r_idx, col_idx, val, fill=cell_fill)

        wb.save(self.output_path)

    # ── Final summary sheet ───────────────────────────────────────────────────

    def write_summary(self, total: int, success: int, failed: int, elapsed_sec: float) -> None:
        # Re-open the file we just flushed and add the summary sheet
        wb = openpyxl.load_workbook(self.output_path)

        if "Automation_Summary" in wb.sheetnames:
            del wb["Automation_Summary"]
        ws2 = wb.create_sheet("Automation_Summary")

        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 18

        headers = [("Metric", "Value")]
        rows = [
            ("Total Records",       str(total)),
            ("Successful",          str(success)),
            ("Failed",              str(failed)),
            ("Success Rate",        f"{success / total * 100:.1f}%" if total else "0%"),
            ("Execution Time (s)",  f"{elapsed_sec:.1f}"),
            ("Completed At",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        _hdr_cell(ws2, 1, 1, "Metric", bg=SUMMARY_HDR_BG)
        _hdr_cell(ws2, 1, 2, "Value",  bg=SUMMARY_HDR_BG)

        for r_idx, (metric, value) in enumerate(rows, 2):
            ws2.cell(row=r_idx, column=1, value=metric).font = Font(
                name=FONT_NAME, size=FONT_SIZE, bold=True)
            ws2.cell(row=r_idx, column=2, value=value).font = Font(
                name=FONT_NAME, size=FONT_SIZE)
            for col in (1, 2):
                ws2.cell(row=r_idx, column=col).border = _BORDER
                ws2.cell(row=r_idx, column=col).alignment = Alignment(vertical="center")

        # Colour the success/failed counts
        ws2.cell(row=3, column=2).fill = PatternFill("solid", start_color="C6EFCE")
        ws2.cell(row=4, column=2).fill = PatternFill("solid", start_color="FFC7CE")

        ws2.row_dimensions[1].height = 24

        wb.save(self.output_path)
